from selenium import webdriver
from time import sleep
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import os
import pathlib
import openpyxl
from openpyxl import load_workbook

class TinoMstCrawler:
    def __init__(self,minpage,maxpage,url,filename):
        self.minpage = minpage
        self.maxpage = maxpage
        self.domain = 'https://masothue.com'
        self.url = url
        self.bot = webdriver.Chrome()
        self.filename = filename

    def crawl(self):
        bot = self.bot
        for i in range(self.minpage,self.maxpage):
            bot.get(self.domain+self.url+'?page='+str(i))
            print('Open browser' + self.domain+self.url + 'with page - ' + str(i))
            time.sleep(3)
            elements = bot.find_elements(By.XPATH, '//div[@class = "tax-listing"]//div[@data-prefetch != ""]//h3//a')
            urls = set()
            for elem in elements:
                urls.add(elem.get_attribute('href'))

            for url in urls:
                bot.get(url)
                time.sleep(3)
                try: 
                    tax_id = bot.find_element(By.XPATH, '//table[@class="table-taxinfo"]//td[@itemprop="taxID"]//span').text
                except:
                    tax_id = ''
                    time.sleep(30)
                try: 
                    name = bot.find_element(By.XPATH, '//table[@class = "table-taxinfo"]//th[@itemprop="name"]/span').text
                except:
                    name = ''
                try: 
                    phone = bot.find_element(By.XPATH, '//table[@class = "table-taxinfo"]//td[@itemprop="telephone"]/span').text
                except:
                    phone = ''

                wb = openpyxl.load_workbook('congty.xlsx')
                ws = wb.active
                max_row = ws.max_row

                ws[f'A{max_row + 1}'] = max_row
                ws[f'B{max_row + 1}'] = tax_id
                ws[f'C{max_row + 1}'] = name
                ws[f'D{max_row + 1}'] = phone
                print('Save excel ' + tax_id + ' | ' + name + ' | ' + phone )
                wb.close()
                wb.save('congty.xlsx')
                time.sleep(3)

# config minpage - maxxpage - Link - FileName

tino = TinoMstCrawler(25,30,'/tra-cuu-ma-so-thue-doanh-nghiep-moi-thanh-lap/','congty.xlsx')
tino.crawl()


#name = bot.find_element(By.XPATH, '//table[@class = "table-taxinfo"]//th[@itemprop="name"]/span').text
# international_name = bot.find_elements(By.XPATH, '//table[@class = "table-taxinfo"]//i[contains(@class, "fa-globe")]/parent::td/following-sibling::td[@itemprop="alternateName"]/span/text()')
# short_name = bot.find_elements(By.XPATH, '//table[@class = "table-taxinfo"]//i[contains(@class, "fa-reorder")]/parent::td/following-sibling::td[@itemprop="alternateName"]/span/text()')
# representative = bot.find_elements(By.XPATH, '//table[@class = "table-taxinfo"]//td/span[@itemprop="name"]/a/text()')
# company_type = bot.find_elements(By.XPATH, '//table[@class = "table-taxinfo"]//td/i[contains(@class, "fa-building")]/parent::td/following-sibling::td/a/text()')
# industry = bot.find_elements(By.XPATH, '//h3[contains(text(), "Ngành nghề kinh doanh")]//following-sibling::table//td/strong/a/text()')
# address = bot.find_elements(By.XPATH, '//table[@class = "table-taxinfo"]//td[@itemprop="address"]/span/text()')
# phone = bot.find_element(By.XPATH, '//table[@class = "table-taxinfo"]//td[@itemprop="telephone"]/span').text
# active_date = bot.find_elements(By.XPATH, '//table[@class = "table-taxinfo"]//td/i[contains(@class, "fa-calendar")]/parent::td/following-sibling::td/span/text()')
# status = bot.find_elements(By.XPATH, '//table[@class = "table-taxinfo"]//td/i[contains(@class, "fa-info")]/parent::td/following-sibling::td/a/text()')
# last_update = bot.find_elements(By.XPATH, '//table[@class = "table-taxinfo"]//button[@data-target = "#modal-update"]/preceding-sibling::em/text()')
# print('Get thông tin thành công')   
